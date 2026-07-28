#!/usr/bin/env python3
"""NHT CLD — ERC-matching Excel layout"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

PI = dict(
    doc_title='COMPLEX LOOP DIAGRAM WITH NARRATIVES', unit='UNIT 400 - NHT',
    proj_no='GP0901', proj_name='MRC CCR + HDS COMPLEX PROJECT',
    company='MIDLAND REFINERIES COMPANY (MRC)', contractor='GPEM',
    site='DAURA REFINERY, IRAQ', doc_no='TBD', rev='0A', date='28 Jul 2026',
)
REVS = [('0A','28 Jul 2026','ISSUED FOR REVIEW','ACB','---','---')]

LOOPS = [
    dict(
        tab_name='NHT-LD-001', loop_id='400-PIC-005',
        loop_title='FEED SURGE DRUM\nPRESSURE SPLIT-RANGE CONTROL', pid_no='TBD',
        desc=(
            '1.  400-PIC-005 CONTROLS THE VAPOUR SPACE PRESSURE OF FEED SURGE DRUM 400-D-401.\n\n'
            '2.  CONTROLLER ACTION: REVERSE ACTION — PV RISES, OUTPUT DECREASES.\n\n'
            '3.  NORMAL OPERATING PRESSURE: REFER TO P&ID / DATA SHEET.\n\n'
            '4.  LOW PRESSURE (OP 50–100%): MAKEUP GAS VALVE 400-PV-005A OPENS PROGRESSIVELY '
            'TO ADMIT H2 RECYCLE GAS FROM MAKEUP GAS SUPPLY HEADER. FAIL POSITION: AO/AFC.\n\n'
            '5.  HIGH PRESSURE (OP 0–50%): VENT VALVE 400-PV-005B OPENS PROGRESSIVELY TO '
            'RELEASE GAS TO FUEL GAS HEADER / FLARE. FAIL POSITION: AO/AFC.\n\n'
            '6.  SPLIT-RANGE COMPUTED BY DCS ARTHC FUNCTION BLOCKS 400-PY-005A (0–50%) '
            'AND 400-PY-005B (50–100%).'
        ),
        comp=(
            'OP   0 – 50%   :  PY-005A  IN 0%->50%    OUT 100%->0%   ->  PV-005B  (VENT, AO/AFC)\n'
            'OP  50–100%  :  PY-005B  IN 50%->100%  OUT 0%->100%  ->  PV-005A  (MAKEUP GAS, AO/AFC)'
        ),
        png='/tmp/cld_work/loop1.png',
    ),
    dict(
        tab_name='NHT-LD-002', loop_id='400-TIC-020 / 400-FC-022',
        loop_title='STRIPPER OVERHEAD TEMPERATURE\nCASCADE TO REFLUX FLOW', pid_no='TBD',
        desc=(
            '1.  400-TIC-020 CONTROLS OVERHEAD TEMPERATURE OF NAPHTHA STRIPPER 400-C-401 BY '
            'CASCADING ITS OUTPUT AS REMOTE SETPOINT TO REFLUX FLOW CONTROLLER 400-FC-022.\n\n'
            '2.  PRIMARY (MASTER) 400-TIC-020: REVERSE ACTION. PV = OVERHEAD TEMPERATURE VIA '
            '400-TT-020. RISING TEMPERATURE INCREASES OUTPUT, DEMANDING MORE REFLUX.\n\n'
            '3.  SECONDARY (SLAVE) 400-FC-022: REVERSE ACTION. PV = REFLUX FLOW VIA 400-FT-022. '
            'SP = REMOTE SETPOINT FROM 400-TIC-020 WHEN IN CASCADE MODE.\n\n'
            '4.  FALLING REFLUX FLOW INCREASES 400-FC-022 OUTPUT, OPENING 400-FCV-022 (AO/AFC).\n\n'
            '5.  THE FAST INNER FLOW LOOP REJECTS HYDRAULIC DISTURBANCES BEFORE THEY PROPAGATE '
            'TO THE SLOWER OUTER TEMPERATURE LOOP.\n\n'
            '6.  BUMPLESS TRANSFER LOGIC ENSURES SMOOTH HANDOVER BETWEEN CASCADE AND MANUAL MODES.'
        ),
        comp=(
            'MASTER  :  TIC-020   (PV: OVERHEAD TEMP — TT-020)\n'
            '                  |\n'
            '              REMOTE SP  (CASCADE MODE)\n'
            '                  |\n'
            'SLAVE   :  FC-022    (PV: REFLUX FLOW — FT-022)\n'
            '                  |\n'
            'VALVE   :  FCV-022   AO / AFC'
        ),
        png='/tmp/cld_work/loop2.png',
    ),
]
OUTPUT = '/tmp/cld_work/NHT_CLD_Sample.xlsx'

NAVY='1F3864'; STEEL='2E4D7B'; LGREY='D9D9D9'; OFFWH='F2F2F2'; WHITE='FFFFFF'; BLACK='000000'

def F(h):  return PatternFill('solid', fgColor=h)
def Ft(name='Arial Narrow',sz=8,bold=False,color=BLACK,italic=False):
    return Font(name=name,size=sz,bold=bold,color=color,italic=italic)
def Al(h='left',v='center',wrap=True,rot=0):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap,text_rotation=rot)
def Sd(s=None,c=BLACK): return Side(style=s,color=c) if s else Side(style=None)
def Bd(t=None,b=None,l=None,r=None):
    return Border(top=Sd(t),bottom=Sd(b),left=Sd(l),right=Sd(r))

THIN=Bd('thin','thin','thin','thin')
MED =Bd('medium','medium','medium','medium')

def mw(ws,r1,c1,r2,c2,val='',fill=None,fnt=None,aln=None,bdr=None):
    if not (r1==r2 and c1==c2):
        ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    cell=ws.cell(r1,c1)
    cell.value=val
    if fill: cell.fill=fill
    if fnt:  cell.font=fnt
    if aln:  cell.alignment=aln
    if bdr:  cell.border=bdr
    return cell

def hline(ws,row,c1,c2,s='thin'):
    sd=Sd(s)
    for c in range(c1,c2+1):
        ex=ws.cell(row,c).border
        ws.cell(row,c).border=Border(top=ex.top,bottom=sd,left=ex.left,right=ex.right)

def vline(ws,r1,r2,col,s='thin'):
    sd=Sd(s)
    for r in range(r1,r2+1):
        ex=ws.cell(r,col).border
        ws.cell(r,col).border=Border(top=ex.top,bottom=ex.bottom,left=sd,right=ex.right)

def box(ws,r1,c1,r2,c2,s='medium'):
    sd=Sd(s); n=Sd(None)
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            t=sd if r==r1 else n; b=sd if r==r2 else n
            l=sd if c==c1 else n; rg=sd if c==c2 else n
            ex=ws.cell(r,c).border
            ws.cell(r,c).border=Border(
                top   =t  if t.style  else ex.top,
                bottom=b  if b.style  else ex.bottom,
                left  =l  if l.style  else ex.left,
                right =rg if rg.style else ex.right)

def a4l(ws,nc,nr):
    ws.page_setup.orientation='landscape'; ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.page_setup.fitToPage=False; ws.page_setup.scale=100
    m=0.315
    ws.page_margins.left=ws.page_margins.right=m
    ws.page_margins.top=ws.page_margins.bottom=m
    ws.page_margins.header=ws.page_margins.footer=0.0
    ws.print_area=f'A1:{get_column_letter(nc)}{nr}'
    ws.sheet_view.showGridLines=False

# ── Grid ─────────────────────────────────────────────────────────────────────
NC=28; NR=50; RH=11
DIAG_C1=1;  DIAG_C2=17   # A-Q diagram
RP_C1=18;   RP_C2=28     # R-AB right panel
NR_MAIN=42               # main rows
BS1=43; BS2=50           # bottom strip rows

# Bottom strip column sections
BRT1=1;  BRT2=10         # revision table
BLG1=11; BLG2=14         # logos
BID1=15; BID2=20         # loop id + title
BTB1=21; BTB2=28         # title block

# Right panel rows
RCD=(1,2); RDESC=(3,24); RCP=(25,26); RCOMP=(27,42)


def _rev_table(ws, r1, r2, c1, c2, hdrs, revs):
    """Build ERC-style revision table: column headers at bottom, data above."""
    tu=sum(u for _,u in hdrs); sw=c2-c1+1
    cur=c1; spans=[]
    for lbl,u in hdrs:
        w=max(1,round(u*sw/tu)); spans.append((cur,cur+w-1,lbl)); cur+=w
    spans[-1]=(spans[-1][0],c2,spans[-1][2])

    # Column header — bottom row
    for cs,ce,lbl in spans:
        mw(ws,r2,cs,r2,ce,lbl,
           F(LGREY),Ft(sz=6.5,bold=True),Al('center','center',False),THIN)

    # Data rows: 2 rows each, placed immediately above column header
    for i,rev in enumerate(revs):
        dr=r2-2-i*2
        if dr < r1: break
        for (cs,ce,_),val in zip(spans,rev):
            mw(ws,dr,cs,dr+1,ce,val,
               F(OFFWH),Ft(sz=6.5),Al('center','center',False),THIN)

    # Blank rows at top (future revisions)
    blank_end=r2-2-len(revs)*2
    if blank_end >= r1:
        # fill white and add outer box
        for r in range(r1,blank_end+1):
            for c in range(c1,c2+1):
                ws.cell(r,c).fill=F(WHITE)
        box(ws,r1,c1,blank_end,c2,'thin')


def build_loop_sheet(wb, loop):
    ws=wb.create_sheet(title=loop['tab_name'])
    a4l(ws,NC,NR)

    for ci in range(DIAG_C1,DIAG_C2+1):
        ws.column_dimensions[get_column_letter(ci)].width=5.5
    for ci in range(RP_C1,RP_C2+1):
        ws.column_dimensions[get_column_letter(ci)].width=4.5
    for ri in range(1,NR+1):
        ws.row_dimensions[ri].height=RH

    # ── DIAGRAM ───────────────────────────────────────────────────────────────
    for r in range(1,NR_MAIN+1):
        for c in range(DIAG_C1,DIAG_C2+1):
            ws.cell(r,c).fill=F(WHITE)
    if os.path.isfile(loop['png']):
        img=XLImage(loop['png'])
        img.width=int(172*3.7796); img.height=int(158*3.7796)
        img.anchor='A1'; ws.add_image(img)
    box(ws,1,DIAG_C1,NR_MAIN,DIAG_C2,'medium')

    # ── RIGHT PANEL ───────────────────────────────────────────────────────────
    S,E=RP_C1,RP_C2
    for r in range(1,NR_MAIN+1):
        for c in range(S,E+1):
            ws.cell(r,c).fill=F(WHITE)

    def rp(r1,r2,val='',fill=None,fnt=None,aln=None):
        mw(ws,r1,S,r2,E,val,fill or F(WHITE),fnt,aln)

    rp(*RCD,'CONTROL DESCRIPTION',F(NAVY),Ft(sz=8.5,bold=True,color=WHITE),Al('left','center',False))
    rp(*RDESC,loop['desc'],F(WHITE),Ft(sz=7),Al('left','top',True))
    rp(*RCP,'COMPUTATION',F(STEEL),Ft(sz=8.5,bold=True,color=WHITE),Al('left','center',False))
    rp(*RCOMP,loop['comp'],F(WHITE),Ft(sz=7,name='Courier New'),Al('left','top',True))

    hline(ws,RCD[1],S,E,'thin')
    hline(ws,RDESC[1],S,E,'thin')
    hline(ws,RCP[1],S,E,'thin')
    box(ws,1,S,NR_MAIN,E,'medium')

    # ── BOTTOM STRIP (rows 43-50, full width) ─────────────────────────────────
    for r in range(BS1,BS2+1):
        for c in range(1,NC+1):
            ws.cell(r,c).fill=F(WHITE)

    # 1. Revision table
    REV_H=[('REV.',1),('DATE',2),('DESCRIPTION',4),('DRAWN',1),('CHKD',1),('APPD',1)]
    _rev_table(ws,BS1,BS2,BRT1,BRT2,REV_H,REVS)

    # 2. Logos — left half GPEM, right half MRC
    LM=BLG1+(BLG2-BLG1)//2; LM2=LM+1
    mw(ws,BS1,BLG1,BS2,LM,PI['contractor'],
       F(WHITE),Ft(sz=9,bold=True,color=NAVY),Al('center','center',False),THIN)
    mw(ws,BS1,LM2,BS2,BLG2,'MRC',
       F(WHITE),Ft(sz=9,bold=True,color=NAVY),Al('center','center',False),THIN)

    # 3. Loop ID + title
    ID_MID=BS1+2
    mw(ws,BS1,BID1,ID_MID,BID2,loop['loop_id'],
       F(WHITE),Ft(sz=9,bold=True),Al('center','center',False))
    hline(ws,ID_MID,BID1,BID2,'thin')
    mw(ws,ID_MID+1,BID1,BS2,BID2,loop['loop_title'],
       F(WHITE),Ft(sz=8,bold=True,color=NAVY),Al('center','center',True))
    box(ws,BS1,BID1,BS2,BID2,'thin')

    # 4. Title block — 4 rows of 2 rows each
    T1=BTB1; T2=BTB2; TM=T1+(T2-T1)//2; TM2=TM+1
    step=2; r=BS1
    mw(ws,r,T1,r+step-1,T2,f"P & ID NO.   :   {loop['pid_no']}",
       F(OFFWH),Ft(sz=6.5),Al('left','center',False)); hline(ws,r+step-1,T1,T2,'thin'); r+=step
    mw(ws,r,T1,r+step-1,T2,f"PROJECT   :   {PI['proj_name']}",
       F(OFFWH),Ft(sz=6.5),Al('left','center',False)); hline(ws,r+step-1,T1,T2,'thin'); r+=step
    mw(ws,r,T1,r+step-1,TM,f"DWG NO.   :   {PI['doc_no']}",
       F(OFFWH),Ft(sz=6.5),Al('left','center',False))
    mw(ws,r,TM2,r+step-1,T2,f"REV.   {PI['rev']}",
       F(OFFWH),Ft(sz=6.5),Al('left','center',False))
    vline(ws,r,r+step-1,TM2,'thin'); hline(ws,r+step-1,T1,T2,'thin'); r+=step
    mw(ws,r,T1,BS2,TM,f"SHEET NO.   :   {loop['tab_name']}",
       F(OFFWH),Ft(sz=6.5),Al('left','center',False))
    mw(ws,r,TM2,BS2,T2,f"JOB NO.   {PI['proj_no']}",
       F(OFFWH),Ft(sz=6.5),Al('left','center',False))
    vline(ws,r,BS2,TM2,'thin')
    box(ws,BS1,T1,BS2,T2,'thin')

    box(ws,BS1,1,BS2,NC,'medium')
    box(ws,1,1,NR,NC,'medium')


def build_blank_sheet(wb,tab_name,title):
    ws=wb.create_sheet(title=tab_name); a4l(ws,NC,NR)
    ws.sheet_view.showGridLines=False
    for ci in range(1,NC+1): ws.column_dimensions[get_column_letter(ci)].width=5.0
    for ri in range(1,NR+1): ws.row_dimensions[ri].height=RH

    mw(ws,1,1,3,NC,f"{PI['contractor']}   |   {PI['doc_title']}   |   {PI['unit']}",
       F(NAVY),Ft(sz=9,bold=True,color=WHITE),Al('center','center',False))
    mw(ws,4,1,8,NC,title,F(WHITE),Ft(sz=14,bold=True,color=NAVY),Al('center','center',False))
    mw(ws,9,1,NR_MAIN,NC,'( THIS SECTION TO BE COMPLETED )',
       F('FAFAFA'),Ft(sz=9,color='BBBBBB',italic=True),Al('center','center',False))

    for r in range(BS1,BS2+1):
        for c in range(1,NC+1): ws.cell(r,c).fill=F(WHITE)

    REV_H=[('REV.',1),('DATE',2),('DESCRIPTION',4),('DRAWN',1),('CHKD',1),('APPD',1)]
    _rev_table(ws,BS1,BS2,BRT1,BRT2,REV_H,REVS)
    LM=BLG1+(BLG2-BLG1)//2
    mw(ws,BS1,BLG1,BS2,LM,PI['contractor'],F(WHITE),Ft(sz=9,bold=True,color=NAVY),Al('center','center',False),THIN)
    mw(ws,BS1,LM+1,BS2,BLG2,'MRC',F(WHITE),Ft(sz=9,bold=True,color=NAVY),Al('center','center',False),THIN)
    mw(ws,BS1,BID1,BS2,BID2,tab_name,F(WHITE),Ft(sz=10,bold=True,color=NAVY),Al('center','center',False),THIN)
    mw(ws,BS1,BTB1,BS2,BTB2,
       f"DOC NO.  :  {PI['doc_no']}\nREV.  {PI['rev']}     DATE:  {PI['date']}",
       F(OFFWH),Ft(sz=6.5),Al('left','center',True),THIN)

    box(ws,BS1,1,BS2,NC,'medium')
    box(ws,1,1,NR,NC,'medium')


def build_cover(wb):
    ws=wb.active; ws.title='COVER'; a4l(ws,NC,NR)
    ws.sheet_view.showGridLines=False
    for ci in range(1,NC+1): ws.column_dimensions[get_column_letter(ci)].width=5.0
    for ri in range(1,NR+1): ws.row_dimensions[ri].height=RH

    mw(ws,1,1,4,NC,f"{PI['doc_title']}\n{PI['unit']}",
       F(NAVY),Ft(sz=12,bold=True,color=WHITE),Al('center','center',True))
    mw(ws,5,1,16,NC,f"{PI['doc_title']}\n\n{PI['unit']}\n\n{PI['proj_name']}",
       F(WHITE),Ft(sz=14,bold=True,color=NAVY),Al('center','center',True))

    sp=7
    for i,(lbl,val) in enumerate([('PROJECT NO.',PI['proj_no']),('PROJECT NAME',PI['proj_name']),
                                   ('CONTRACTOR',PI['contractor']),('COMPANY',PI['company']),
                                   ('SITE',PI['site'])]):
        r=17+i*3
        mw(ws,r,1,r+2,sp,lbl,F(LGREY),Ft(sz=8,bold=True),Al('left','center',False))
        mw(ws,r,sp+1,r+2,NC,f':   {val}',F(WHITE),Ft(sz=9),Al('left','center',False))
        hline(ws,r+2,1,NC,'thin')

    mw(ws,33,7,38,22,'ISSUED FOR REVIEW',
       F(WHITE),Ft(sz=11,bold=True,color='C00000'),Al('center','center',False),MED)
    mw(ws,39,1,NR_MAIN,NC,
       f"DOC NO.  {PI['doc_no']}     |     REV.  {PI['rev']}     |     DATE:  {PI['date']}",
       F(OFFWH),Ft(sz=8),Al('center','center',False))

    for r in range(BS1,BS2+1):
        for c in range(1,NC+1): ws.cell(r,c).fill=F(WHITE)
    half=NC//2
    mw(ws,BS1,1,BS2-2,half,PI['contractor'],F(WHITE),Ft(sz=14,bold=True,color=NAVY),Al('center','center',False),THIN)
    mw(ws,BS1,half+1,BS2-2,NC,'MRC',F(WHITE),Ft(sz=14,bold=True,color=NAVY),Al('center','center',False),THIN)

    REV_H=[('REV.',1),('DATE',2),('DESCRIPTION',4),('PREPARED',1),('CHECKED',1),('APPROVED',1)]
    tu=sum(u for _,u in REV_H); cur=1; spans=[]
    for lbl,u in REV_H:
        w=max(1,round(u*NC/tu)); spans.append((cur,cur+w-1,lbl)); cur+=w
    spans[-1]=(spans[-1][0],NC,spans[-1][2])
    for cs,ce,lbl in spans:
        mw(ws,BS2,cs,BS2,ce,lbl,F(LGREY),Ft(sz=6.5,bold=True),Al('center','center',False),THIN)
    if REVS:
        for (cs,ce,_),val in zip(spans,REVS[0]):
            mw(ws,BS2-1,cs,BS2-1,ce,val,F(OFFWH),Ft(sz=6.5),Al('center','center',False),THIN)

    box(ws,BS1,1,BS2,NC,'medium')
    box(ws,1,1,NR,NC,'medium')


def main():
    wb=Workbook()
    build_cover(wb)
    build_blank_sheet(wb,'NHT-GN-001','GENERAL NOTES')
    build_blank_sheet(wb,'NHT-IN-001','INDEX')
    for loop in LOOPS:
        build_loop_sheet(wb,loop)
    wb.save(OUTPUT)
    print(f'Done: {OUTPUT}')

if __name__=='__main__':
    main()
